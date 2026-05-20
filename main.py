# -*- coding: utf-8 -*-
"""物流服务商对账服务 (WF-C 引擎). 由 n8n 定时 HTTP 触发 /recon/run.
逻辑: 扫物流对账单任务台 触发对账=true -> 下载服务商Excel附件 -> openpyxl解析按运单号聚合
 -> join 物流对账明细(服务商运单号) -> 我方侧按配置公式(关联发货任务.总计+明细字段) 逐票算
 -> 与服务商侧逐票比 -> 回写汇总/差异/结果/快照+关联明细 -> 飞书卡片通知物流(DRY_RUN只发Frankie)."""
import os, io, json, time, urllib.request, urllib.error
from fastapi import FastAPI, Header, HTTPException
from openpyxl import load_workbook

app = FastAPI()

E = os.environ.get
APP1 = (E("FEISHU_APP1_ID", ""), E("FEISHU_APP1_SECRET", ""))
APP3 = (E("FEISHU_APP3_ID", ""), E("FEISHU_APP3_SECRET", ""))
BT = E("BITABLE_APP_TOKEN", "")
T_RECON = E("TBL_RECON", "")
T_DETAIL = E("TBL_DETAIL", "")
T_TASK = E("TBL_TASK", "")
T_CFG = E("TBL_CFG", "")
FRANKIE = E("FRANKIE_OID", "")
LOGI_DEPT = E("LOGI_DEPT_ID", "")
LOGI_JOB = E("LOGI_JOB", "物流仓储主管")
DRY_RUN = E("DRY_RUN", "1") == "1"
BEARER = E("BEARER", "")
FB = "https://open.feishu.cn"
T_BILL = E("TBL_BILL", "tblF4e0SfakHGLS8")
T_SUPPLIER_CFG = E("TBL_SUPPLIER_CFG", "")
T_DECL = E("TBL_DECL", "")
LINGXING_PROXY_URL = E("LINGXING_PROXY_URL", "")
PROXY_TOKEN = E("PROXY_TOKEN", "")

# 店铺+国家 → 领星 sid 映射 (procurement memory 同步)
SHOP_SID = {
    ("Fanlepu", "美国"): "3841", ("Fanlepu", "加拿大"): "3842", ("Fanlepu", "墨西哥"): "3843",
    ("FUNLAB", "美国"): "1182", ("FUNLAB", "加拿大"): "1197", ("FUNLAB", "日本"): "1200",
    ("FUNLAB", "墨西哥"): "2650",
    ("Funlab Collection", "澳洲"): "1198", ("Funlab Collection", "澳大利亚"): "1198",
    ("FunlabDirect", "英国"): "1192", ("FunlabDirect", "德国"): "1194",
    ("FunlabDirect", "法国"): "1195", ("FunlabDirect", "西班牙"): "1196",
    ("FunlabDirect", "意大利"): "1193",
    ("pl-us", "美国"): "3621", ("pl-us", "加拿大"): "3622", ("pl-us", "墨西哥"): "3623",
    ("YKTRICH", "美国"): "4339", ("YKTRICH", "加拿大"): "4340", ("YKTRICH", "墨西哥"): "4341",
}


def shop_to_sid(shop, country):
    if not shop:
        return None
    # 精确匹配
    if (shop, country) in SHOP_SID:
        return SHOP_SID[(shop, country)]
    # 前缀模糊匹配 (店铺名可能含后缀)
    for (s, c), sid in SHOP_SID.items():
        if shop.startswith(s) and country == c:
            return sid
    return None


def _req(url, method="GET", body=None, headers=None, raw=False, timeout=120):
    data = None
    if body is not None:
        data = body if isinstance(body, (bytes, bytearray)) else json.dumps(body).encode("utf-8")
    h = {} if raw else {"Content-Type": "application/json"}
    if headers:
        h.update(headers)
    rq = urllib.request.Request(url, data=data, headers=h, method=method)
    try:
        with urllib.request.urlopen(rq, timeout=timeout) as r:
            c = r.read()
            return c if raw else json.loads(c.decode("utf-8"))
    except urllib.error.HTTPError as e:
        c = e.read()
        return c if raw else json.loads(c.decode("utf-8"))


def tok(app):
    r = _req(FB + "/open-apis/auth/v3/tenant_access_token/internal", "POST",
             {"app_id": app[0], "app_secret": app[1]})
    return r.get("tenant_access_token", "")


def nz(v):
    if v is None:
        return ""
    if isinstance(v, list):
        return "".join((x.get("text", "") if isinstance(x, dict) else str(x)) for x in v)
    if isinstance(v, dict):
        return v.get("text", "")
    return v


def num(v):
    s = nz(v)
    try:
        return float(str(s).replace(",", "").strip() or 0)
    except Exception:
        return 0.0


def lk(v):
    if not v:
        return []
    if isinstance(v, list):
        o = []
        for x in v:
            if isinstance(x, str):
                o.append(x)
            elif isinstance(x, dict):
                if x.get("record_ids"):
                    o += x["record_ids"]
                elif x.get("id"):
                    o.append(x["id"])
        return o
    if isinstance(v, dict):
        return v.get("link_record_ids") or v.get("record_ids") or []
    return []


def list_records(t1, table):
    out, pt = [], ""
    while True:
        u = (FB + "/open-apis/bitable/v1/apps/%s/tables/%s/records?page_size=200" % (BT, table)
             + (("&page_token=" + pt) if pt else ""))
        r = _req(u, "GET", None, {"Authorization": "Bearer " + t1})
        d = r.get("data", {}) or {}
        for it in d.get("items", []):
            out.append({"rid": it["record_id"], "f": it.get("fields", {})})
        if d.get("has_more") and d.get("page_token"):
            pt = d["page_token"]
        else:
            break
    return out


def batch_get(t1, table, ids):
    if not ids:
        return {}
    r = _req(FB + "/open-apis/bitable/v1/apps/%s/tables/%s/records/batch_get" % (BT, table),
             "POST", {"record_ids": ids}, {"Authorization": "Bearer " + t1})
    res = {}
    for x in (r.get("data", {}) or {}).get("records", []):
        res[x["record_id"]] = x.get("fields", {})
    return res


def put_rec(t1, table, rid, fields):
    return _req(FB + "/open-apis/bitable/v1/apps/%s/tables/%s/records/%s" % (BT, table, rid),
                "PUT", {"fields": fields}, {"Authorization": "Bearer " + t1})


def resolve_logi(t1):
    """物流仓储部 LOGI_JOB 职务 在职 open_id (铁律:职务实时查,不硬编码)."""
    out, pt = [], ""
    while True:
        u = (FB + "/open-apis/contact/v3/users?department_id=%s&department_id_type=department_id"
             "&user_id_type=open_id&page_size=50" % LOGI_DEPT + (("&page_token=" + pt) if pt else ""))
        r = _req(u, "GET", None, {"Authorization": "Bearer " + t1})
        d = r.get("data", {}) or {}
        for usr in d.get("items", []):
            act = (usr.get("status") or {}).get("is_activated", True)
            if act and usr.get("job_title") == LOGI_JOB:
                out.append((usr.get("open_id"), usr.get("name")))
        if d.get("has_more") and d.get("page_token"):
            pt = d["page_token"]
        else:
            break
    return out


def send_card(t3, oid, title, lines, warn):
    els = [{"tag": "div", "text": {"tag": "lark_md",
            "content": "\n".join(lines)}}]
    card = {"config": {"wide_screen_mode": True},
            "header": {"template": "red" if warn else "turquoise",
                       "title": {"tag": "plain_text", "content": title}},
            "elements": els}
    return _req(FB + "/open-apis/im/v1/messages?receive_id_type=open_id", "POST",
                {"receive_id": oid, "msg_type": "interactive",
                 "content": json.dumps(card, ensure_ascii=False)},
                {"Authorization": "Bearer " + t3})


def download_media(t1, file_token):
    return _req(FB + "/open-apis/drive/v1/medias/%s/download" % file_token, "GET", None,
                {"Authorization": "Bearer " + t1}, raw=True)


def parse_supplier_excel(content, cfg):
    """按配置解析 xlsx, 返回 {运单号: {amount: sum折合人民币, fees:[(费用类型,金额)]}}."""
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    hr = int(cfg["hr"])
    ds = int(cfg["ds"])
    rows = list(ws.iter_rows(values_only=True))
    if hr - 1 >= len(rows):
        return {}, []
    header = ["" if c is None else str(c).strip() for c in rows[hr - 1]]
    jc = cfg["jcol"]
    ac = cfg["acol"]
    fc = cfg.get("fcol", "")
    ji = header.index(jc) if jc in header else -1
    ai = header.index(ac) if ac in header else -1
    fi = header.index(fc) if fc in header else -1
    if ji < 0 or ai < 0:
        return {}, ["表头缺列: join=%s amount=%s header=%s" % (jc, ac, header)]
    agg = {}
    warns = []
    for r in rows[ds - 1:]:
        if r is None:
            continue
        key = "" if ji >= len(r) or r[ji] is None else str(r[ji]).strip()
        if not key:
            continue
        amt = 0.0
        if ai < len(r) and r[ai] is not None:
            try:
                amt = float(str(r[ai]).replace(",", "").strip() or 0)
            except Exception:
                amt = 0.0
        ft = ""
        if 0 <= fi < len(r) and r[fi] is not None:
            ft = str(r[fi]).strip()
        a = agg.setdefault(key, {"amount": 0.0, "fees": []})
        a["amount"] += amt
        if ft:
            a["fees"].append((ft, amt))
    return agg, warns


@app.get("/health")
def health():
    return {"ok": True, "dry_run": DRY_RUN, "service": "recon-wf-c"}


@app.post("/recon/run")
def recon_run(authorization: str = Header(default="")):
    if BEARER and authorization != "Bearer " + BEARER:
        raise HTTPException(status_code=401, detail="unauthorized")
    t1 = tok(APP1)
    t3 = tok(APP3) if APP3[0] else t1
    if not t1:
        raise HTTPException(status_code=500, detail="feishu token fail")

    # 1. 服务商映射配置
    cfgmap = {}
    for it in list_records(t1, T_CFG):
        f = it["f"]
        sup = nz(f.get("服务商"))
        if not sup:
            continue
        en = f.get("启用")
        if en is False:
            continue
        cfgmap[sup] = {
            "hr": int(num(f.get("Excel表头行")) or 1),
            "ds": int(num(f.get("Excel数据起始行")) or 2),
            "jcol": nz(f.get("Excel_join列名")),
            "acol": nz(f.get("Excel_金额列名")),
            "fcol": nz(f.get("Excel_费用类型列名")),
            "myjoin": nz(f.get("我方join字段")),
            "formula": nz(f.get("我方汇总公式")),
            "thr": num(f.get("差异阈值百分比")) or 5.0,
        }

    # 2. 待对账的对账单 (触发对账=true)
    recons = [it for it in list_records(t1, T_RECON) if it["f"].get("触发对账") is True]
    logi = resolve_logi(t1)
    logi_oids = [o for o, _ in logi] or []
    logi_names = "、".join(n for _, n in logi) or "-"

    # 3. 预读全部物流对账明细 (含关联发货任务 -> 取 TASK 字段)
    details = list_records(t1, T_DETAIL)
    task_ids = []
    for d in details:
        task_ids += lk(d["f"].get("关联发货任务"))
    tasks = batch_get(t1, T_TASK, list(set(task_ids))) if task_ids else {}

    out = []
    for rec in recons:
        rid = rec["rid"]
        f = rec["f"]
        sup = nz(f.get("服务商"))
        c = cfgmap.get(sup)
        if not c:
            put_rec(t1, T_RECON, rid, {"对账明细快照": "无映射配置, 请先在『服务商对账映射配置』加 %s 并启用" % sup,
                                       "触发对账": False})
            put_rec(t1, T_RECON, rid, {"对账结果": "待对账"})
            out.append({"rid": rid, "sup": sup, "act": "no-config"})
            continue
        atts = f.get("服务商对账单附件") or []
        ft = atts[0].get("file_token") if atts and isinstance(atts[0], dict) else None
        if not ft:
            put_rec(t1, T_RECON, rid, {"对账明细快照": "未上传服务商对账单附件", "触发对账": False})
            out.append({"rid": rid, "sup": sup, "act": "no-attach"})
            continue
        content = download_media(t1, ft)
        try:
            agg, warns = parse_supplier_excel(content, c)
        except Exception as ex:
            put_rec(t1, T_RECON, rid, {"对账明细快照": "Excel解析失败: %s" % str(ex)[:200],
                                       "触发对账": False})
            out.append({"rid": rid, "sup": sup, "act": "parse-fail", "err": str(ex)[:200]})
            continue

        # 4. join 我方物流对账明细 (服务商运单号 == Excel join key)
        toks = [x.strip() for x in str(c["formula"]).split("+") if x.strip()]
        sup_total = sum(v["amount"] for v in agg.values())
        my_total = 0.0
        matched_dids = []
        snap = []
        n_diff = 0
        n_match = 0
        for wb_no, sv in agg.items():
            md = None
            for d in details:
                if nz(d["f"].get(c["myjoin"])).strip() == wb_no:
                    md = d
                    break
            if md is None:
                snap.append("%s 服务商¥%.2f 我方<无明细>" % (wb_no, sv["amount"]))
                continue
            n_match += 1
            matched_dids.append(md["rid"])
            myv = 0.0
            tf = {}
            tlinks = lk(md["f"].get("关联发货任务"))
            if tlinks:
                tf = tasks.get(tlinks[0], {})
            for tk in toks:
                if "." in tk:
                    sc, fld = tk.split(".", 1)
                    if sc == "TASK":
                        myv += num(tf.get(fld))
                    elif sc == "DETAIL":
                        myv += num(md["f"].get(fld))
            my_total += myv
            diff = myv - sv["amount"]
            pct = (abs(diff) / sv["amount"] * 100) if sv["amount"] else (100 if myv else 0)
            ok = pct <= c["thr"]
            if not ok:
                n_diff += 1
            snap.append("%s 服务商¥%.2f 我方¥%.2f 差%.2f(%.1f%%) %s"
                        % (wb_no, sv["amount"], myv, diff, pct, "平" if ok else "差异"))
            put_rec(t1, T_DETAIL, md["rid"],
                    {"对账状态": "对账平" if ok else "对账有差异"})

        result = "对账平" if (n_diff == 0 and n_match == len(agg) and n_match > 0) else (
            "有差异" if n_diff > 0 else "部分匹配")
        diff_total = my_total - sup_total
        pct_total = (abs(diff_total) / sup_total * 100) if sup_total else 0
        snaptxt = ("服务商账单¥%.2f / 我方汇总¥%.2f / 差异¥%.2f (%.1f%%) / 命中%d票 / 服务商总票%d\n"
                   % (sup_total, my_total, diff_total, pct_total, n_match, len(agg))
                   ) + "\n".join(snap[:60])
        put_rec(t1, T_RECON, rid, {
            "服务商账单金额": round(sup_total, 2),
            "我方汇总金额": round(my_total, 2),
            "命中票数": n_match,
            "差异金额": round(diff_total, 2),
            "差异百分比": round(pct_total, 2),
            "对账明细快照": snaptxt[:9000],
            "关联物流对账明细": matched_dids,
            "已对账": True,
            "触发对账": False,
        })
        put_rec(t1, T_RECON, rid, {"对账结果": result})

        warn = (result != "对账平")
        period = ""
        try:
            ps, pe = f.get("对账周期起"), f.get("对账周期止")
            period = "%s ~ %s" % (
                time.strftime("%Y-%m-%d", time.gmtime(ps / 1000)) if ps else "?",
                time.strftime("%Y-%m-%d", time.gmtime(pe / 1000)) if pe else "?")
        except Exception:
            period = "-"
        lines = [
            "**服务商**：%s" % sup,
            "**对账周期**：%s" % period,
            "**服务商账单**：¥%.2f" % sup_total,
            "**我方汇总**：¥%.2f（命中 %d / 服务商 %d 票）" % (my_total, n_match, len(agg)),
            "**差异**：¥%.2f（%.1f%%）" % (diff_total, pct_total),
            "**结果**：%s" % result,
            "差异票/快照见对账单『对账明细快照』，请核对后改『处理状态』。",
        ]
        if DRY_RUN:
            lines = ["⚠ **DRY-RUN** 真实应发：物流仓储主管（%s）" % logi_names] + lines
            targets = [FRANKIE]
        else:
            targets = logi_oids or [FRANKIE]
        mids = []
        for o in targets:
            # 用聪哥1号发: resolve_logi/FRANKIE 的 open_id 均属聪哥1号命名空间
            # (聪哥3号发会 99992361 open_id cross app)
            rr = send_card(t1, o, "物流对账 · %s %s" % (sup, result), lines, warn)
            mids.append((rr.get("data", {}) or {}).get("message_id")
                        or ("ERR:" + json.dumps(rr, ensure_ascii=False)[:120]))
        out.append({"rid": rid, "sup": sup, "result": result, "sup_total": round(sup_total, 2),
                    "my_total": round(my_total, 2), "matched": n_match, "diff_rows": n_diff,
                    "mids": mids})

    return {"dry_run": DRY_RUN, "processed": len(out), "detail": out}


def lingxing_call(method, path, params=None):
    """走 lingxing-proxy on n8n (Zeabur 出站 IP 已白名单 + 内置签名)."""
    if not LINGXING_PROXY_URL or not PROXY_TOKEN:
        return {"code": -1, "message": "lingxing proxy not configured"}
    body = {"method": method, "path": path, "params": params or {}}
    return _req(LINGXING_PROXY_URL, "POST", body,
                {"Authorization": "Bearer " + PROXY_TOKEN}, timeout=60)


def lingxing_listings_by_sid(sid):
    """拉某 sid 全部 listings (分页)."""
    out, offset = [], 0
    while True:
        r = lingxing_call("POST", "/erp/sc/data/mws/listing",
                          {"sid": sid, "length": 200, "offset": offset})
        d = r.get("data") or []
        if not d:
            break
        out.extend(d)
        if len(d) < 200:
            break
        offset += 200
        if offset >= 2000:
            break
    return out


def lingxing_all_products():
    """拉全部本地产品 (分页)."""
    out, offset = [], 0
    while True:
        r = lingxing_call("POST", "/erp/sc/routing/data/local_inventory/productList",
                          {"length": 200, "offset": offset})
        d = r.get("data") or []
        if not d:
            break
        out.extend(d)
        if len(d) < 200:
            break
        offset += 200
        if offset >= 2000:
            break
    return out


def bitable_find(t1, table, key_field, key_value):
    """飞书 bitable 按字段值精确匹配查 1 条 (in-mem filter)."""
    if not key_value:
        return None
    for it in list_records(t1, table):
        if nz(it["f"].get(key_field)).strip() == str(key_value).strip():
            return it["f"]
    return None


@app.post("/h4/auto-fill")
def h4_auto_fill(payload: dict, authorization: str = Header(default="")):
    """H4 自动建开票要求明细骨架. WF-A 在主表推到「已发货待开票要求」时调用.
    数据流: task → 领星查 supplier+采购价 → 飞书合规表 + 报关表 → 建明细(状态=待财务核对)."""
    if BEARER and authorization != "Bearer " + BEARER:
        raise HTTPException(status_code=401, detail="unauthorized")
    task_id = (payload or {}).get("task_record_id")
    if not task_id:
        raise HTTPException(status_code=400, detail="missing task_record_id")
    t1 = tok(APP1)
    if not t1:
        raise HTTPException(status_code=500, detail="feishu token fail")

    # 1. 读 task
    r = _req(FB + "/open-apis/bitable/v1/apps/%s/tables/%s/records/%s" % (BT, T_TASK, task_id),
             "GET", None, {"Authorization": "Bearer " + t1})
    tf = (r.get("data", {}) or {}).get("record", {}).get("fields", {})
    if not tf:
        return {"error": "task_not_found", "task_id": task_id}
    if lk(tf.get("关联开票要求")):
        return {"skip": "already_has_bill", "task_id": task_id}

    fnsku = nz(tf.get("FNSKU")).strip()
    pn = nz(tf.get("品名")).strip()
    qty = num(tf.get("发货数量"))
    ship_date = tf.get("实际发货时间") or 0
    sn = nz(tf.get("货件编号")).strip()
    shop = nz(tf.get("店铺")).strip()
    country = nz(tf.get("国家")).strip()
    if not fnsku or not pn:
        return {"error": "missing_fnsku_or_pn", "task_id": task_id}

    # 2. 领星: FNSKU → listing → local_sku → product → supplier/cg_price
    sid = shop_to_sid(shop, country)
    notes = ["H4自动建"]
    if not sid:
        return {"error": "no_sid_mapping", "shop": shop, "country": country,
                "hint": "在 main.py SHOP_SID 加映射"}
    listings = lingxing_listings_by_sid(sid)
    listing = next((l for l in listings if (l.get("fnsku") or "").strip() == fnsku), None)
    if not listing:
        return {"error": "fnsku_not_in_listings", "fnsku": fnsku, "sid": sid,
                "listing_count": len(listings)}
    local_sku = (listing.get("local_sku") or "").strip()
    if not local_sku:
        return {"error": "no_local_sku", "fnsku": fnsku}

    products = lingxing_all_products()
    product = next((p for p in products if (p.get("sku") or "").strip() == local_sku), None)
    if not product:
        return {"error": "sku_not_in_products", "local_sku": local_sku}

    cg_price = num(product.get("cg_price"))
    supplier_name = ""
    supplier_quote = product.get("supplier_quote") or []
    if supplier_quote:
        primary = next((q for q in supplier_quote if q.get("is_primary") == 1), supplier_quote[0])
        supplier_name = (primary.get("supplier_name") or "").strip()
        if num(primary.get("cg_price")) > 0:
            cg_price = num(primary.get("cg_price"))
    else:
        notes.append("领星无 supplier_quote, 主供应商待补")

    # 3. 飞书合规表查
    inv_title = supplier_name or ""
    tax_rate = ""
    account_days = 7
    if supplier_name and T_SUPPLIER_CFG:
        cfg_row = bitable_find(t1, T_SUPPLIER_CFG, "供应商名称", supplier_name)
        if cfg_row:
            inv_title = nz(cfg_row.get("开票抬头")).strip() or supplier_name
            tax_rate = nz(cfg_row.get("票面税率")).strip()
            ad = int(num(cfg_row.get("账期天数")) or 0)
            if ad > 0:
                account_days = ad
        else:
            notes.append("供应商[%s]不在合规表" % supplier_name)
    if not tax_rate:
        notes.append("税率合规表未配")

    # 4. 报关表查
    invoice_item_name = pn
    if T_DECL:
        decl_row = bitable_find(t1, T_DECL, "品名", pn)
        if decl_row:
            v = nz(decl_row.get("开票项目名称")).strip()
            if v:
                invoice_item_name = v
        else:
            notes.append("品名[%s]不在报关表, 开票品名用原品名" % pn)

    # 5. 计算
    invoice_amount = round(qty * cg_price, 2)
    bill_deadline = (int(ship_date) if ship_date else int(time.time() * 1000)) + account_days * 86400 * 1000
    bill_id = sn + "-" + pn
    if cg_price <= 0:
        notes.append("采购价为0,请核对")

    # 6. 建开票要求明细
    fields = {
        "开票要求编号": bill_id,
        "关联发货任务": [task_id],
        "开票品名": invoice_item_name,
        "开票金额": invoice_amount,
        "供应商": supplier_name,
        "开票抬头": inv_title,
        "税率": tax_rate,
        "开票时限": bill_deadline,
        "状态": "待财务核对",
        "备注": " | ".join(notes),
    }
    cr = _req(FB + "/open-apis/bitable/v1/apps/%s/tables/%s/records" % (BT, T_BILL),
              "POST", {"fields": fields}, {"Authorization": "Bearer " + t1})
    new_rid = ((cr.get("data") or {}).get("record") or {}).get("record_id")
    return {
        "created": new_rid,
        "task_id": task_id,
        "fields": fields,
        "notes": notes,
        "lingxing_local_sku": local_sku,
        "lingxing_supplier_count": len(supplier_quote),
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.environ.get("PORT", "8080")))
